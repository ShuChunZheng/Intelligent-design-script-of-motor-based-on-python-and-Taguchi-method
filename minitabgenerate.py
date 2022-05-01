# -*- coding:utf-8 -*-
import numpy as numpy
import xlrd
from openpyxl import Workbook

from minitabexecute import extract_data_from_excellist, minitabquit, minitabopgen, dirgenerate, rownumber, \
    rownumbersingle, listcut, deltefile, rownumbersthreerounbo, logdebug

global minitabaddr
global minitbookaddr
global mworkbookaddr
global workbookaddr
global singminitbookaddr
global minitbookroundboaddr

space = ' '
equalestr = '='
ccolumn = 'C'
middle = '-'
semicolon = ';'
periodstr = '.'

lowerisbetter = 'Lower'
higherisbetter = 'Higher'
gsncd = 'Gsn'
gmeanscd = 'Gmeans'


def paradifine():
    global minitabaddr
    global minitbookaddr
    global mworkbookaddr
    global workbookaddr
    global singminitbookaddr
    global minitbookroundboaddr
    diclist = dirgenerate()
    minitabaddr = diclist['minitabaddr']
    minitbookaddr = diclist['minitbookaddr']
    mworkbookaddr = diclist['mworkbookaddr']
    workbookaddr = diclist['workbookaddr']
    singminitbookaddr = diclist['singminitbookaddr']
    minitbookroundboaddr = diclist['minitbookroundboaddr']


def orthogonalcda54(runs111, level1, factors1, lever2, factors2):
    elementlist = [str(runs111), str(level1), str(factors1), str(lever2), str(factors2)]
    elementliststr = ' '.join(elementlist)
    orthogonalcd = 'OADesign'
    orthogonalcda111 = orthogonalcd + space + elementliststr
    return orthogonalcda111


def orthogonalcda(runs11, level1, factors1):
    elementlist = [str(runs11), str(level1), str(factors1)]
    elementliststr = ' '.join(elementlist)
    orthogonalcd = 'Oadesign'
    orthogonalcda11 = orthogonalcd + space + elementliststr
    return orthogonalcda11


def storeccda(firstcc, finalcc):
    stordccd = 'XMatrix'
    storeccda1 = stordccd + space + ccolumn + str(firstcc) + middle + ccolumn + str(finalcc)
    return storeccda1


def levelscda(listlevels=None):
    if listlevels is None:
        listlevels = []
    levelscd = 'Levels'
    listcut(listlevels)
    listlevelsstr = ' '.join(listlevels)
    levelscda1 = levelscd + space + listlevelsstr
    return levelscda1


def columnscda(listcolumns=None):
    if listcolumns is None:
        listcolumns = []
    columnscd = 'Columns'
    listcolumnsstr = ' '.join(listcolumns)
    columnscda1 = columnscd + space + listcolumnsstr
    return columnscda1


def analyzecda(firstcc, finalcc):
    analyzecd = 'Robust'
    analyzecda1 = analyzecd + space + ccolumn + space + equalestr + space + ccolumn + str(
        firstcc) + middle + ccolumn + str(finalcc)
    return analyzecda1


def namecd(factornumber1, factnamelist=None):
    if factnamelist is None:
        factnamelist = []
    namecd1 = 'Name'
    combinelist = []

    if 26 > factornumber1 > 13:
        factnamelistp = ['"' + i + '"' for i in factnamelist]

        clist = ['C' + str(i) for i in range(1, factornumber1 + 2)]
        factnamelistp.insert(0, '"无用参数"')
        for i in range(factornumber1 + 1):
            combinelist.append(clist[i])
            combinelist.append(factnamelistp[i])
    elif factornumber1 >= 26:
        factnamelistp = ['"' + i[0:2] + '"' for i in factnamelist]
        clist = ['C' + str(i) for i in range(1, factornumber1 + 1)]
        for i in range(factornumber1):
            combinelist.append(clist[i])
            combinelist.append(factnamelistp[i])
    else:
        factnamelistp = ['"' + i + '"' for i in factnamelist]
        clist = ['C' + str(i) for i in range(1, factornumber1 + 1)]
        for i in range(factornumber1):
            combinelist.append(clist[i])
            combinelist.append(factnamelistp[i])
    listcut(combinelist)

    combineliststr = ' '.join(combinelist)
    namecda = namecd1 + space + combineliststr + periodstr

    return namecda


def saveasexcel(addr):
    wsavecda = 'WSave ' + '"' + addr + '";' + 'FType;    Excel 97;  Missing;    Numeric \'*\' \'*\';    Text "" "";  ' \
                                              'Replace. '
    return wsavecda


def xlstoxlsx(runs, factornumber, mworkbookaddr_, wbmc=None, elementnumber=None):
    if wbmc is None:
        wbmc = Workbook()
    if elementnumber is None:
        elementnumber = numpy.zeros(factornumber)
    wbminitab = xlrd.open_workbook_xls(mworkbookaddr_)
    wbmcsheet = wbmc['Sheet']
    wbminitabsheet = wbminitab.sheet_by_name('Worksheet 1')
    for i in range(runs + 1):
        for j in range(factornumber):
            try:
                wbmcsheet.cell(i + 5, elementnumber[j] + 2).value = wbminitabsheet.cell(i, j).value
            except IndexError:
                pass


def xlstoxlsx26(runs, factornumber, mworkbookaddr_, wbmc=None, elementnumber=None):
    if wbmc is None:
        wbmc = Workbook()
    if elementnumber is None:
        elementnumber = numpy.zeros(factornumber)
    wbminitab = xlrd.open_workbook_xls(mworkbookaddr_)
    wbmcsheet = wbmc['Sheet']
    wbminitabsheet = wbminitab.sheet_by_name('Worksheet 1')
    for i in range(runs + 1):
        for j in range(1, factornumber + 1):
            try:
                wbmcsheet.cell(i + 5, elementnumber[j - 1] + 2).value = wbminitabsheet.cell(i, j).value
            except IndexError:
                pass


def main(phasecheck=3):
    paradifine()

    if phasecheck == 3:
        rownumber_ = rownumber
        minitbookaddr_ = minitbookaddr
    elif phasecheck == 4:
        rownumber_ = rownumbersthreerounbo
        minitbookaddr_ = minitbookroundboaddr
    else:
        rownumber_ = rownumbersingle
        minitbookaddr_ = singminitbookaddr

    columnnumber1a = 4
    columnnumber1 = 3
    get_num_lista = []
    get_num_list = numpy.zeros((rownumber_, columnnumber1))
    deltefile(workbookaddr)

    wbmc = Workbook()
    wbmcsheet = wbmc['Sheet']
    extract_data_from_excellist(minitbookaddr_, rownumber_, columnnumber1a, 2, 3, 'Sheet', get_num_lista)
    columnnamelist = []
    get_num_listsame = [0 for _ in range(rownumber_)]
    for i in range(0, rownumber_):
        j = i * 4
        if get_num_lista[j + 1] == get_num_lista[j + 2]:
            get_num_listsame[i] = get_num_lista[j + 1]
        else:
            get_num_list[i][0] = get_num_lista[j + 1]
            get_num_list[i][1] = get_num_lista[j + 2]
            get_num_list[i][2] = get_num_lista[j + 3]

        columnnamelist.append(get_num_lista[j])

    elementnumber = numpy.array(numpy.array(numpy.array(get_num_list.sum(1)).nonzero())[0])
    factornumber = elementnumber.shape[0]
    get_num_listreal = []
    columnnamelistreal = []
    if factornumber < 26:
        for i in range(factornumber):
            for j in range(columnnumber1):
                get_num_listreal.append(get_num_list[elementnumber[i]][j])
            columnnamelistreal.append(columnnamelist[elementnumber[i]])
    else:
        for i in range(factornumber):
            for j in range(2):
                get_num_listreal.append(get_num_list[elementnumber[i]][j])
            columnnamelistreal.append(columnnamelist[elementnumber[i]])
    logdebug(factornumber)
    logdebug(get_num_listreal)
    logdebug(columnnamelistreal)

    runs = 0
    deltefile(minitabaddr)
    minitab, mproject = minitabopgen(minitabaddr)
    if factornumber < 1:
        runs = 1

    elif factornumber < 2:
        runs = 3
        for i in range(runs):
            for j in range(rownumber_):
                if get_num_list[j][i] != 0:
                    wbmcsheet.cell(i + 6, j + 2).value = get_num_list[j][i]

    elif factornumber < 5:
        columnlistrange = [str(i) for i in range(1, factornumber + 1)]
        get_num_listrealstr = [str(i) for i in get_num_listreal]
        runs = 9
        mproject.ExecuteCommand(namecd(factornumber, columnnamelistreal))
        mproject.ExecuteCommand(orthogonalcda(runs, 3, factornumber) + semicolon + space + space +
                                storeccda(1, factornumber) + semicolon + space + space +
                                levelscda(get_num_listrealstr) + semicolon + space + space +
                                columnscda(columnlistrange) + periodstr)
        mproject.ExecuteCommand(saveasexcel(mworkbookaddr))
        xlstoxlsx(runs, factornumber, mworkbookaddr, wbmc, elementnumber)

    elif factornumber < 14:
        columnlistrange = [str(i) for i in range(1, factornumber + 1)]
        get_num_listrealstr = [str(i) for i in get_num_listreal]
        runs = 27
        mproject.ExecuteCommand(namecd(factornumber, columnnamelistreal))
        mproject.ExecuteCommand(orthogonalcda(runs, 3, factornumber) + semicolon + space + space +
                                storeccda(1, factornumber) + semicolon + space + space +
                                levelscda(get_num_listrealstr) + semicolon + space + space +
                                columnscda(columnlistrange) + periodstr)
        mproject.ExecuteCommand(saveasexcel(mworkbookaddr))
        xlstoxlsx(runs, factornumber, mworkbookaddr, wbmc, elementnumber)

    elif factornumber < 26:
        columnlistrange = [str(i) for i in range(1, factornumber + 2)]
        get_num_listreal.insert(0, 1)
        get_num_listreal.insert(1, 2)
        get_num_listrealstr = [str(i) for i in get_num_listreal]
        runs = 54
        mproject.ExecuteCommand(namecd(factornumber, columnnamelistreal))
        mproject.ExecuteCommand(orthogonalcda54(runs, 2, 1, 3, factornumber) + semicolon + space + space +
                                storeccda(1, factornumber + 1) + semicolon + space + space +
                                levelscda(get_num_listrealstr) + semicolon + space + space +
                                columnscda(columnlistrange) + periodstr)
        mproject.ExecuteCommand(saveasexcel(mworkbookaddr))
        xlstoxlsx26(runs, factornumber, mworkbookaddr, wbmc, elementnumber)

    elif factornumber <= 31:
        columnlistrange = [str(i) for i in range(1, factornumber + 1)]
        get_num_listrealstr = [str(i) for i in get_num_listreal]
        runs = 32
        mproject.ExecuteCommand(namecd(factornumber, columnnamelistreal))
        mproject.ExecuteCommand(orthogonalcda(runs, 2, factornumber) + semicolon + space + space +
                                storeccda(1, factornumber) + semicolon + space + space +
                                levelscda(get_num_listrealstr) + semicolon + space + space +
                                columnscda(columnlistrange) + periodstr)
        mproject.ExecuteCommand(saveasexcel(mworkbookaddr))
        xlstoxlsx(runs, factornumber, mworkbookaddr, wbmc, elementnumber)

    mproject.SaveAs(minitabaddr)
    for i in range(runs):
        for j in range(rownumber_):
            if get_num_listsame[j] != 0:
                wbmcsheet.cell(i + 6, j + 2).value = get_num_listsame[j]

    wbmc.save(workbookaddr)
    minitabquit(minitab)


if __name__ == '__main__':
    main(4)
