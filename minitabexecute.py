# -*- coding:utf-8 -*-
# from __future__ import print_function
import ctypes
import logging
import os
import random
import sqlite3
import subprocess
import time
import zlib
from zipfile import BadZipFile

import numpy
import pythoncom
import win32com.client
import wmi
from openpyxl import load_workbook, Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from win32com.universal import com_error

global minitabaddr
global minitbookaddr
global workbookaddr
global plotorkbookaddr
global singminitbookaddr
global logautorunaddr
global minitbookroundboaddr

space = ' '
equalestr = '='
ccolumn = 'C'
middle = '-'
semicolon = ';'
lowerisbetter = 'Lower'
higherisbetter = 'Higher'
gsncd = 'Gsn'
gmeanscd = 'Gmeans'
periodstr = '.'
plotmap = 'GSN;  Gmeans;  TSN;  Tmeans.'
plotmapnt = 'GSN;  Gmeans.'
rownumber = 24
rownumbersingle = 31
rownumbersthreerounbo = 30

rownumbermax = 54
namelist = ['匝数1', '转速1', '铁长0.1', '铁壳厚0.1', '齿宽0.1', '气隙0.01', '槽满率0.01', 'Br 系数0.1', '极弧系数1',
            '超前角0.1', '电压0.1', '峰值电流0.1', '磁铁厚度0.1', '定子外径0.1', '槽口宽0.1', '槽顶深0.1', '槽底圆角0.1',
            '槽顶角度1', '定子扭角0.1', '转子扭角0.1', '定子铁耗系数0.1', '转子铁耗系数0.1', '槽深0.1', '固定转子外径0.1', '0',
            '效率', '齿槽转矩波动', '转矩', '转矩波动', '铜耗', '定子铁耗', '转子铁耗', '线端电压峰值', '母线平均电流', '转子轭磁密', '漆包线直径']

namelistsingle = ['匝数1', '转速1', '铁长0.1', '铁壳厚0.1', '齿宽0.1', '气隙0.01', '槽满率0.01', 'Br 系数0.1', '极弧系数1',
                  '超前角0.1', '电压0.1', '峰值电流0.1', '磁铁厚度0.1', '定子外径0.1', '槽口宽0.1', '槽顶深0.1', '槽底圆角0.1',
                  '槽尖圆角0.1', '定子扭角0.1', '转子扭角0.1', '定子铁耗系数0.1', '转子铁耗系数0.1', '槽深0.1', '固定转子外径0.1',
                  '定子内径0.1', '极对数1', '齿顶厚度中系数0.1', '齿顶厚度大系数0.1', '外径下沉大0.01', '外径下沉中系数0.01', '槽顶圆角0.1', '0',
                  '效率', '齿槽转矩波动', '转矩', '转矩波动', '铜耗', '定子铁耗', '转子铁耗', '线端电压峰值', '母线平均电流', '转子轭磁密', '漆包线直径']

namelistthreer = ['匝数1', '转速1', '铁长0.1', '铁壳厚0.1', '齿宽0.1', '气隙0.01', '槽满率0.01', 'Br 系数0.1', '极弧系数1',
                  '超前角0.1', '电压0.1', '峰值电流0.1', '磁铁厚度0.1', '定子外径0.1', '槽口宽0.1', '槽顶深0.1', '槽底圆角0.1',
                  '槽尖圆角0.1', '定子扭角0.1', '转子扭角0.1', '定子铁耗系数0.1', '转子铁耗系数0.1', '槽深0.1', '固定转子外径0.1',
                  '定子内径0.1', '极对数1', '齿顶厚度中系数0.1', '齿顶厚度大系数0.1', '槽顶圆角0.1', '槽数*3 1', '0',
                  '效率', '齿槽转矩波动', '转矩', '转矩波动', '铜耗', '定子铁耗', '转子铁耗', '线端电压峰值', '母线平均电流', '转子轭磁密', '漆包线直径']

"""
Created on Mon Jan  8 09:09:51 2018

@author: coordinate
"""

quickdatanam = 'quickdata'
fulldatanam = 'fulldatanam'
returnrown = 10
phasechecktabstr = 'phasecheck'
numbercheckstr = 'numbercheck'
quickcalcheckstr = 'quickcalcheck'
timeratiocheckstr = 'timeratiocheck'
threetimeratiocheckstr = 'threetimeratiocheck'
finishcheckstr = 'finishcheck'
motorinpufincheckstr = 'motorinpufincheck'
motorinputablenamestr = 'motorinputablename'
motorinputreducesigetstr = 'motorinputreducesiget'
turnscaccheckstr = 'turnscaccheck'
turnscachigeffcheckstr = 'turnscachigeffcheck'


def dirgenerate():
    pythoncom.CoInitialize()
    setupdir_ = app_path() + r'\lib'
    pythontrialexcel = setupdir_ + r'\pythontrialexcel'
    plotorkbookaddr_ = setupdir_ + r'\performancecurve' + '.xlsx'
    plotorkbookaddr_tem = setupdir_ + r'\performancecurvetem' + '.xlsx'
    workbookaddr_ = pythontrialexcel + '.xlsx'
    MotorCAD_Fileshot = setupdir_ + r'\pythontrial'
    minitbookaddr_ = setupdir_ + r'\minitabtrial' + '.xlsx'
    minitbookaddr_temp = setupdir_ + r'\minitabtrial' + '_temp' + '.xlsx'

    minitbookroundboaddr_ = setupdir_ + r'\minitabtrialroundbo' + '.xlsx'
    minitbookroundboaddr_temp = setupdir_ + r'\minitabtrialroundbo' + '_temp' + '.xlsx'

    pyseworkbookaddr1 = pythontrialexcel + '1.xlsx'
    pyseworkbookaddr2 = pythontrialexcel + '2.xlsx'
    pyseworkbookaddr3 = pythontrialexcel + '3.xlsx'
    pyseworkbookaddr4 = pythontrialexcel + '4.xlsx'

    MotorCAD_File1 = MotorCAD_Fileshot + '_' + '1.mot'
    MotorCAD_File2 = MotorCAD_Fileshot + '_' + '2.mot'
    MotorCAD_File3 = MotorCAD_Fileshot + '_' + '3.mot'
    MotorCAD_File4 = MotorCAD_Fileshot + '_' + '4.mot'
    MotorCAD_File5 = MotorCAD_Fileshot + '_' + '5.mot'

    workbookaddr1 = pythontrialexcel + '1' + '.xlsx'
    workbookaddr2 = pythontrialexcel + '2' + '.xlsx'
    workbookaddr3 = pythontrialexcel + '3' + '.xlsx'
    workbookaddr4 = pythontrialexcel + '4' + '.xlsx'
    minitabaddr_ = setupdir_ + r'\Minitab.MPJ'
    minitabaddr_tem = setupdir_ + r'\Minitabtem.MPJ'
    mworkbookaddr = setupdir_ + r'\minitabtrialws.xls'
    modeladdr = setupdir_ + r'\modellib'
    singlepmodeladdr = modeladdr + r'\singlephase'
    threepmodeladdr = modeladdr + r'\threephase'
    threepmodelroundboaddr = modeladdr + r'\threephaseroundbo'
    testresultaddr = modeladdr + r'\testresult' + '.xlsx'
    testresultafteraddr = modeladdr + r'\testresultafter' + '.xlsx'
    higheffresultfoldaddr = modeladdr + r'\higheffresultfold'
    singminitbookaddr_ = setupdir_ + r'\singminitabtrial' + '.xlsx'
    singminitbookaddr_temp = setupdir_ + r'\singminitabtrial' + '_temp' + '.xlsx'
    dexfileloca = setupdir_ + r'\dxf.dxf'
    dexfileloca1 = setupdir_ + r'\dxf1.dxf'
    dexfileloca2 = setupdir_ + r'\dxf2.dxf'
    dexfileloca3 = setupdir_ + r'\dxf3.dxf'
    dexfileloca4 = setupdir_ + r'\dxf4.dxf'
    bigfiled1 = MotorCAD_Fileshot + r'_1'
    bigfiled2 = MotorCAD_Fileshot + r'_2'
    bigfiled3 = MotorCAD_Fileshot + r'_3'
    bigfiled4 = MotorCAD_Fileshot + r'_4'
    logautorunaddr_ = app_path() + r'\log_autorun.log'
    logautorunaddrtem = app_path() + r'\log_autoruntem.log'
    logmotorcad1addr = app_path() + r'\log_motorcad1.log'
    logmotorcad2addr = app_path() + r'\log_motorcad2.log'
    logmotorcad3addr = app_path() + r'\log_motorcad3.log'
    logmotorcad4addr = app_path() + r'\log_motorcad4.log'
    logmotorcadinputaddr = app_path() + r'\log_motorcadinput.log'

    return {'setupdir': setupdir_, 'pythontrialexcel': pythontrialexcel, 'threepmodeladdr': threepmodeladdr,
            'bigfiled3': bigfiled3, 'logautorunaddr': logautorunaddr_, 'pyseworkbookaddr2': pyseworkbookaddr2,
            'pyseworkbookaddr3': pyseworkbookaddr3, 'minitabaddr_tem': minitabaddr_tem,
            'threepmodelroundboaddr': threepmodelroundboaddr,
            'logmotorcad2addr': logmotorcad2addr, 'logmotorcadinputaddr': logmotorcadinputaddr,
            'plotorkbookaddr_tem': plotorkbookaddr_tem, 'logmotorcad1addr': logmotorcad1addr,
            'logmotorcad3addr': logmotorcad3addr, 'minitbookroundboaddr': minitbookroundboaddr_,
            'minitbookroundboaddr_temp': minitbookroundboaddr_temp,
            'plotorkbookaddr': plotorkbookaddr_, 'workbookaddr': workbookaddr_, 'singlepmodeladdr': singlepmodeladdr,
            'MotorCAD_Fileshot': MotorCAD_Fileshot, 'minitbookaddr': minitbookaddr_, 'dexfileloca1': dexfileloca1,
            'singminitbookaddr': singminitbookaddr_, 'dexfileloca3': dexfileloca3, 'dexfileloca2': dexfileloca2,
            'pyseworkbookaddr1': pyseworkbookaddr1, 'minitbookaddr_temp': minitbookaddr_temp, 'bigfiled1': bigfiled1,
            'singminitbookaddr_temp': singminitbookaddr_temp, 'dexfileloca': dexfileloca, 'dexfileloca4': dexfileloca4,
            'MotorCAD_File1': MotorCAD_File1, 'workbookaddr1': workbookaddr1, 'workbookaddr2': workbookaddr2,
            'workbookaddr3': workbookaddr3, 'workbookaddr4': workbookaddr4, 'minitabaddr': minitabaddr_,
            'bigfiled4': bigfiled4, 'MotorCAD_File2': MotorCAD_File2, 'MotorCAD_File3': MotorCAD_File3,
            'MotorCAD_File4': MotorCAD_File4, 'logautorunaddrtem': logautorunaddrtem,
            'logmotorcad4addr': logmotorcad4addr, 'MotorCAD_File5': MotorCAD_File5, 'higheffresultfoldaddr': higheffresultfoldaddr,
            'mworkbookaddr': mworkbookaddr, 'bigfiled2': bigfiled2, 'modeladdr': modeladdr,
            'pyseworkbookaddr4': pyseworkbookaddr4, 'testresultaddr': testresultaddr, 'testresultafteraddr':testresultafteraddr}


def extract_data_from_excel_after(addr_, rownumber_, rownumberexcel, rownumber1, rownumberall, columnnumber21,
                                  num_list0=None):
    if num_list0 is None:
        num_list0 = numpy.zeros((rownumberall, columnnumber21))

    while 1:
        try:
            wb11 = load_workbook(addr_)
            for j in range(columnnumber21):
                row1 = rownumberexcel + 6
                column1 = j + rownumber_ + 3
                num_list0[rownumber1][j] = wb11['Sheet'].cell(row=row1, column=column1).value
                if numpy.isnan(num_list0[rownumber1][j]):
                    num_list0[rownumber1][j] = 0
                if num_list0[rownumber1][j] < 0:
                    num_list0[rownumber1][j] = 0
            time.sleep(0.01)
            wb11.close()
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def extract_data_from_excel(addr_, rownumber0, columnnumber0, joffset, num_list0=None):
    rownumberture = rownumber0
    if num_list0 is None:
        num_list0 = numpy.zeros((rownumber0, columnnumber0))
    while 1:
        try:
            wb11 = load_workbook(addr_)
            for i in range(rownumber0):
                zerocheck = 0
                for j in range(columnnumber0):
                    row1 = i + 6
                    column1 = j + joffset
                    num_list0[i][j] = wb11['Sheet'].cell(row=row1, column=column1).value
                    if numpy.isnan(num_list0[i][j]):
                        num_list0[i][j] = 0
                    zerocheck += num_list0[i][j]
                if zerocheck == 0:
                    rownumberture = i
                    break
            time.sleep(0.1)
            wb11.close()
            return rownumberture
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def logconfi(logautorunaddr_):
    logging.basicConfig(filename=logautorunaddr_, level=logging.INFO)


def logconfidebug(logautorunaddr_):
    logging.basicConfig(filename=logautorunaddr_, level=logging.DEBUG)


def loggenera(message):
    logging.info(message)


def logdebug(message):
    logging.debug(message)


def deltefile(dexfileloca_):
    while 1:
        try:
            os.remove(dexfileloca_)
            break
        except FileNotFoundError:
            break
        except PermissionError:
            time.sleep(1)
            continue


def copyfile(souraddr, tageaddr):
    while 1:
        try:
            commandstrminitb = 'copy %s %s' % (souraddr, tageaddr)
            os.system(commandstrminitb)
            break
        except FileNotFoundError:
            break
        except PermissionError:
            time.sleep(1)
            continue


def bounddry(centre1, numblist=None):
    if numblist is None:
        numblist = []
    maxc = max(numblist)
    minc = min(numblist)
    if centre1 >= maxc:
        x1 = minc
        x2 = centre1 + centre1 - minc
    elif centre1 <= minc:
        x1 = centre1 + centre1 - maxc
        x2 = maxc
    else:
        if centre1 - minc >= maxc - centre1:
            x1 = minc
            x2 = centre1 + centre1 - minc
        else:
            x1 = centre1 + centre1 - maxc
            x2 = maxc
    return x1, x2


def funcflatpara(centre):
    a = 100
    d = centre
    b = (a - 0.01) / d ** 4
    return [a, d, b]


def fuctionflat(x, funcflatp=None):
    if funcflatp is None:
        funcflatp = []
    a = funcflatp[0]
    d = funcflatp[1]
    b = funcflatp[2]
    b1 = (a - 0.01) * 7.673 * 10 ** (-6) / d ** 4
    if x <= d:
        y = round(a - b * (x - d) ** 4, 2)
    else:
        y = round(a - b1 * (x - d) ** 4, 2)
    if y <= 0:
        y = 0.001
    return y


def fuctionflatc(x, funcflatp=None):
    if funcflatp is None:
        funcflatp = []
    a = funcflatp[0]
    b = 0
    d = 0
    if x > funcflatp[1]:
        d = funcflatp[1]
        b = funcflatp[2]
    y = round(a - b * (x - d) ** 4, 2)
    if y <= 0:
        y = 0.001
    return y


def funfinal(signal, x, y, z):
    funfinalv = 0
    if signal == 0:
        funfinalv = round(x * y * z / 1111, 0)
    if signal == 1:
        funfinalv = round(x * y * z / 1111, 0)
    if funfinalv <= 0:
        funfinalv = 0.0001
    return funfinalv


def creattable(sqlitedbn, sqlitetablename, columnpl, connet=None):
    if connet is None:
        connet = sqlite3.connect(sqlitedbn + '.db')
    columnnamelist = ['D{} REAL'.format(str(i)) for i in range(columnpl + 1)]
    if columnpl == 0:
        columnname = columnnamelist[0]
    else:
        columnname = ', '.join(columnnamelist)
    createcomm = "CREATE TABLE IF NOT EXISTS %s(%s);" % (sqlitetablename, columnname)
    connet.execute(createcomm)
    connet.commit()
    connet.close()


def insertdata(rownumbercol, sqlitedbn, sqlitetablename, columnpl, array=None, connet=None):
    if connet is None:
        connet = sqlite3.connect(sqlitedbn + '.db')
    if array is None:
        array = numpy.empty((rownumbercol, columnpl + 1))

    templist = ['?' for _ in range(columnpl + 1)]
    if columnpl == 0:
        symbolstr = templist[0]
    else:
        symbolstr = ', '.join(templist)
    insertcommstr = "INSERT INTO %s VALUES " % sqlitetablename
    insertcommtem = insertcommstr + "(%s);" % symbolstr
    connet.executemany(insertcommtem, array)
    connet.commit()
    connet.close()


def selectdata(sqlitedbn, sqlitetablename, returnrown_, columnpl, connet=None):
    if connet is None:
        connet = sqlite3.connect(sqlitedbn + '.db')
    try:
        selectcommstr = "SELECT * FROM %s WHERE D%d <= 900 ORDER BY D%d DESC LIMIT %d" % (
            sqlitetablename, columnpl, columnpl, returnrown_)
        cursor = connet.execute(selectcommstr)
        array = numpy.zeros((returnrown_, columnpl + 2))
        i = 0
        for row in cursor:
            for j in range(columnpl + 2):
                if j == columnpl + 1:
                    array[i][j] = i
                else:
                    array[i][j] = row[j]
            i = i + 1
        connet.close()
        return array

    except sqlite3.OperationalError:
        logdebug(sqlite3.OperationalError)
        pass


def plothodata(plotorkbookaddr_, sqlitetablename, returnrown_, columnpl, startrow, sheet='Sheet1', dbname=fulldatanam):
    arrayreturn = selectdata(dbname, sqlitetablename, returnrown_, columnpl)
    write_data_to_excelhol(plotorkbookaddr_, returnrown_, columnpl + 2, startrow, 1, sheet, arrayreturn)


def numberinput(dbname, sqlitetablename, dataarraynum):
    try:
        dataarray = numpy.empty((1, 1))
        dataarray[0][0] = dataarraynum
        creattable(dbname, sqlitetablename, 0)
        insertdata(1, dbname, sqlitetablename, 0, dataarray)
    except Exception:
        logdebug(Exception)


def numberget(dbname, sqlitetablename):
    try:
        connet = sqlite3.connect(dbname + '.db')
        selectcommstr = "SELECT * FROM %s" % sqlitetablename
        cursor = connet.execute(selectcommstr)
        checknumb = 0
        for row in cursor:
            checknumb = row[0]
        connet.close()
        return checknumb
    except sqlite3.OperationalError:
        loggenera(sqlite3.OperationalError)
        pass


def mainsqlite(rownumber_, minitbookaddr__, workbookaddr_, rownumbercol, sqlitetablename, signal=0, dbname=fulldatanam):
    columnplo = rownumber_ + 12
    dataarray = numpy.zeros((rownumbercol, columnplo + 1))
    extract_data_from_excelarray(workbookaddr_, rownumbercol, columnplo, 6, 2, 'Sheet', dataarray)
    efficiencylist = []
    cogginglist = []
    torquelist = []
    for i in range(rownumbercol):
        efficiencylist.append(dataarray[i][rownumber_ + 1])
        cogginglist.append(dataarray[i][rownumber_ + 2])
        torquelist.append(dataarray[i][rownumber_ + 3])
    currentlist = []
    extract_data_from_excellist(minitbookaddr__, 2, 1, 3, 9, 'Sheet', currentlist)
    paralistt = funcflatpara(currentlist[0])
    paralistc = []
    if currentlist[1] == 0 and currentlist[0] != 0:
        paralistc = funcflatpara(currentlist[0] * 0.3)

    elif currentlist[0] != 0 and currentlist[1] != 0:
        paralistc = funcflatpara(currentlist[1])

    for i in range(rownumbercol):
        dataarray[i][columnplo] = funfinal(signal, efficiencylist[i], fuctionflat(torquelist[i], paralistt),
                                           fuctionflatc(cogginglist[i], paralistc))
    creattable(dbname, sqlitetablename, columnplo)
    insertdata(rownumbercol, dbname, sqlitetablename, columnplo, dataarray)


def inputfilehan(minitbookaddr_, minitbookaddr_temp_, threepmodeladdr_, MotorCAD_Fileshot_, namelist_=None):
    if namelist_ is None:
        namelist_ = []
    namestrinlist = filename(threepmodeladdr_)
    namestrinfirst = ','.join(namestrinlist)
    namestrin = '"%s"' % namestrinfirst
    creatminitexcel(minitbookaddr_temp_, namestrin, namelist_)
    exchangeexcel(minitbookaddr_, minitbookaddr_temp_, 2, namelist_.index('0') + 1, 2, 2)
    exchangeexcel(minitbookaddr_, minitbookaddr_temp_, 2, namelist_.index('0') + 1, 7, 8)
    exchangeexcel(minitbookaddr_, minitbookaddr_temp_, 2, 11, 9, 9)
    exchangeexcel(minitbookaddr_, minitbookaddr_temp_, 17, 17, 10, 10)
    exchangeexcel(minitbookaddr_, minitbookaddr_temp_, 21, 21, 10, 10)
    exchangeexcel(minitbookaddr_, minitbookaddr_temp_, 24, 24, 10, 10)

    deltefile(minitbookaddr_)
    copyfile(minitbookaddr_temp_, minitbookaddr_)
    modelname = getdata(minitbookaddr_, 17, 10)
    MotorCAD_Fileshothol = MotorCAD_Fileshot_ + '.mot'
    deltefile(MotorCAD_Fileshothol)
    copyfile(threepmodeladdr_ + r'\%s.mot' % modelname, MotorCAD_Fileshothol)


def openexcel(filenamem):
    while 1:
        try:
            xlapp = win32com.client.Dispatch("Excel.Application")
            xlapp.Visible = True
            xlapp.Workbooks.Open(filenamem)
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def filename(modeladdr):
    modnamelist = os.listdir(modeladdr)
    modnamelistture = []
    for i in range(len(modnamelist)):
        if modnamelist[i].endswith('.mot'):
            modnamelistture.append(modnamelist[i].replace('.mot', ''))
    return modnamelistture


def pythoncomundo():
    pythoncom.CoUninitialize()


def getsysteminf():
    cpulist = wmi.WMI().Win32_Processor()
    for i in cpulist:
        loggenera(i)


def killprocess():
    process_name = ['RuntimeBroker.exe', 'Mtb.exe', 'Motor-CAD_13_1_10.exe', 'dispatcher.exe', 'OSPPSVC.EXE',
                    'MsMpEng.exe', 'SWVisualize.BoostService.exe', 'SWVisualize.Queue.Server.exe']
    devnull = open(os.devnull, 'w')
    for i in process_name:
        commandstr = 'taskkill /IM %s /F' % i
        subprocess.Popen(commandstr, stdout=devnull, stderr=devnull)

def killexcellprocess():
    process_name = ['EXCEL.exe']
    devnull = open(os.devnull, 'w')
    for i in process_name:
        commandstr = 'taskkill /IM %s /F' % i
        subprocess.Popen(commandstr, stdout=devnull, stderr=devnull)


def killprocessall():
    process_name = ['RuntimeBroker.exe', 'Mtb.exe', 'Motor-CAD_13_1_10.exe', 'dispatcher.exe', 'OSPPSVC.EXE', 'cmd.exe',
                    'conhost.exe', 'MsMpEng.exe', 'SWVisualize.BoostService.exe', 'SWVisualize.Queue.Server.exe',
                    'python.exe', 'pythonw.exe']
    devnull = open(os.devnull, 'w')
    for i in process_name:
        commandstr = 'taskkill /IM %s /F' % i
        subprocess.Popen(commandstr, stdout=devnull, stderr=devnull)


def quickedit(enabled=1):  # This is a patch to the system that sometimes hangs
    """
    Enable or disable quick edit mode to prevent system hangs, sometimes when using remote desktop
    Param (Enabled)
    enabled = 1(default), enable quick edit mode in python console
    enabled = 0, disable quick edit mode in python console
    """
    # -10 is input handle => STD_INPUT_HANDLE (DWORD) -10 |
    # https://docs.microsoft.com/en-us/windows/console/getstdhandle default = (0x4|0x80|0x20|0x2|0x10|0x1|0x40|0x200)
    # 0x40 is quick edit, #0x20 is insert mode 0x8 is disabled by default
    # https://docs.microsoft.com/en-us/windows/console/setconsolemode
    kernel32 = ctypes.windll.kernel32
    if enabled:
        kernel32.SetConsoleMode(kernel32.GetStdHandle(-10), (0x4 | 0x80 | 0x20 | 0x2 | 0x10 | 0x1 | 0x40 | 0x100))
    else:
        kernel32.SetConsoleMode(kernel32.GetStdHandle(-10), (0x4 | 0x80 | 0x20 | 0x2 | 0x10 | 0x1 | 0x00 | 0x100))


def runfatorn(minitbookaddr_, rownumber_):
    get_num_lista = []
    extract_data_from_excellist(minitbookaddr_, rownumber_, 1, 2, 2, 'Sheet', get_num_lista)
    factornumber = 0
    limitarray = numpy.zeros((rownumber_, 2))
    extract_data_from_excelarray(minitbookaddr_, rownumber_, 2, 2, 7, 'Sheet', limitarray)
    for i in range(rownumber_):
        if limitarray[i][0] == limitarray[i][1] != 0:
            get_num_lista[i] = 0
    for i in range(rownumber_):
        if get_num_lista[i] != 0:
            factornumber = factornumber + 1
    runs = 0
    if factornumber < 1:
        runs = 1
    elif factornumber < 2:
        runs = 3
    elif factornumber < 5:
        runs = 9
    elif factornumber < 14:
        runs = 27
    elif factornumber < 26:
        runs = 54
    elif 26 <= factornumber <= 31:
        runs = 32
    return runs, factornumber, get_num_lista


def paradifine():
    global minitabaddr
    global minitbookaddr
    global workbookaddr
    global plotorkbookaddr
    global singminitbookaddr
    global logautorunaddr
    global minitbookroundboaddr
    diclist = dirgenerate()
    minitabaddr = diclist['minitabaddr']
    minitbookaddr = diclist['minitbookaddr']
    workbookaddr = diclist['workbookaddr']
    plotorkbookaddr = diclist['plotorkbookaddr']
    singminitbookaddr = diclist['singminitbookaddr']
    logautorunaddr = diclist['logautorunaddr']
    minitbookroundboaddr = diclist['minitbookroundboaddr']


def app_path():
    return os.path.dirname(__file__)


def orthogonalcda(runs, level1, factors1):
    elementlist = [str(runs), str(level1), str(factors1)]
    elementliststr = ' '.join(elementlist)
    orthogonalcd = 'Oadesign'
    orthogonalcda1 = orthogonalcd + space + elementliststr
    return orthogonalcda1


def storeccda(firstcc, finalcc):
    stordccd = 'XMatrix'
    storeccda1 = stordccd + space + ccolumn + str(firstcc) + middle + ccolumn + str(finalcc)
    return storeccda1


def designcda(firstcc, finalcc):
    designccd = 'Design'
    designccda = designccd + space + ccolumn + str(firstcc) + middle + ccolumn + str(finalcc)
    return designccda


def levelscda(listlevels=None):
    if listlevels is None:
        listlevels = []
    levelscd = 'Levels'
    intesp = 25
    intdiv = len(listlevels) // intesp
    if intdiv >= 2:
        for i in range(1, intdiv):
            listlevels.insert(i * intesp - 1, '&')

    listlevelsstr = ' '.join(listlevels)
    levelscda1 = levelscd + space + listlevelsstr
    return levelscda1


def listcut(listlevels=None):
    if listlevels is None:
        listlevels = []
    intesp = 15
    intdiv = len(listlevels) // intesp
    if intdiv >= 2:
        for i in range(1, intdiv):
            listlevels.insert(i * intesp - 1, "&\n")


def columnscda(listcolumns=None):
    if listcolumns is None:
        listcolumns = []
    columnscd = 'Columns'
    listcolumnsstr = ' '.join(listcolumns)
    columnscda1 = columnscd + space + listcolumnsstr
    return columnscda1


def analyzecda(analysecc, firstcc, finalcc):
    analyzecd = 'Robust'
    analyzecda1 = analyzecd + space + ccolumn + str(analysecc) + space + equalestr + space + ccolumn + str(
        firstcc) + middle + ccolumn + str(finalcc)
    return analyzecda1


def maxindex(factornumber1, stringtxt, max_number=None):
    if max_number is None:
        max_number = numpy.zeros(factornumber1)

    strlist = stringtxt.split()
    liststr1 = []
    liststr2 = []
    liststr3 = []

    for i1 in range(len(strlist)):
        if strlist[i1] == 'Delta':
            for j1 in range(i1, len(strlist)):
                if strlist[j1] == 'Level':
                    break
                strlist[j1] = 'NONE'
    for i1 in range(len(strlist)):
        if strlist[i1] == '1':
            for j1 in range(i1 + 1, len(strlist)):
                if strlist[j1] == '2':
                    for d in range(j1 + 1, len(strlist)):
                        if strlist[d] == '3':
                            for e in range(d + 1, len(strlist)):
                                if strlist[e] == 'NONE':
                                    break
                                liststr3.append(strlist[e])
                            break
                        liststr2.append(strlist[d])
                    break
                liststr1.append(strlist[j1])
    listnumber1 = [eval(i1) for i1 in liststr1]
    listnumber2 = [eval(i1) for i1 in liststr2]
    listnumber3 = [eval(i1) for i1 in liststr3]
    while len(listnumber1) < factornumber1:
        listnumber1.append(0)

    while len(listnumber2) < factornumber1:
        listnumber2.append(0)

    while len(listnumber3) < factornumber1:
        listnumber3.append(0)

    max_index = numpy.zeros(factornumber1)
    for i1 in range(factornumber1):
        listnumberall = [listnumber1[i1], listnumber2[i1], listnumber3[i1]]
        max_number[i1] = max(listnumberall)
        max_index[i1] = listnumberall.index(max_number[i1])

    return max_index


def maxindextwolev(factornumber1, stringtxt, max_number=None):
    if max_number is None:
        max_number = numpy.zeros(factornumber1)

    strlist = stringtxt.split()
    liststr1 = []
    liststr2 = []

    for i1 in range(len(strlist)):
        if strlist[i1] == 'Delta':
            for j1 in range(i1, len(strlist)):
                if strlist[j1] == 'Level':
                    break
                strlist[j1] = 'NONE'
    for i1 in range(len(strlist)):
        if strlist[i1] == '1':
            for j1 in range(i1 + 1, len(strlist)):
                if strlist[j1] == '2':
                    for d in range(j1 + 1, len(strlist)):
                        if strlist[d] == 'NONE':
                            break
                        liststr2.append(strlist[d])
                    break
                liststr1.append(strlist[j1])

    listnumber1 = [eval(i1) for i1 in liststr1]
    listnumber2 = [eval(i1) for i1 in liststr2]
    while len(listnumber1) < factornumber1:
        listnumber1.append(0)

    while len(listnumber2) < factornumber1:
        listnumber2.append(0)

    max_index = numpy.zeros(factornumber1)
    for i1 in range(factornumber1):
        listnumberall = [listnumber1[i1], listnumber2[i1]]
        max_number[i1] = max(listnumberall)
        max_index[i1] = listnumberall.index(max_number[i1])

    return max_index


def writemaxindex(minitbookaddr__, rownumber_, factornumber1, maxindexarray=None, get_num_lista21=None):
    if maxindexarray is None:
        maxindexarray = numpy.zeros(factornumber1)
    if get_num_lista21 is None:
        get_num_lista21 = [0 for _ in range(rownumber_)]
    listtemple = [0 * i1 for i1 in range(rownumber_)]
    while 1:
        try:
            wb = load_workbook(minitbookaddr__)
            maxindexarrayin = 0
            for i1 in range(rownumber_):
                if get_num_lista21[i1] != 0:
                    listtemple[i1] = wb['Sheet'].cell(row=i1 + 2,
                                                      column=maxindexarray[maxindexarrayin] + 4).value
                    maxindexarrayin = maxindexarrayin + 1
            time.sleep(0.1)
            wb.close()
            return listtemple
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def extractdatathreeroundbignu(minitbookaddr__, signal, factornumber1, rownumber_, mproject1=None,
                               get_num_lista21=None):
    if mproject1 is None:
        mproject1 = win32com.client.Dispatch('{D3407EDD-1BAE-4E00-A512-2673EBF849CE}').ActiveProject
    if get_num_lista21 is None:
        get_num_lista21 = [0 for _ in range(rownumber_)]
    max_number = numpy.zeros(factornumber1)
    listtemple = []
    for j1 in range(1, mproject1.Commands.Count + 1):
        command = mproject1.Commands(j1)
        signals = 0
        for e in range(1, command.Outputs.Count + 1):
            number1 = command.Outputs(e).OutputType
            if number1 == 1:
                responsestr = str(command.Outputs(e).Table.Text)
                if 'Response Table for Means' in responsestr:
                    loggenera('均值相应表：' + '\n' + responsestr)
                    maxindexarray = maxindextwolev(factornumber1, responsestr, max_number)
                    listtemple = writemaxindex(minitbookaddr__, rownumber_, factornumber1, maxindexarray,
                                               get_num_lista21)
                    signals = 1
                    break
        if signals == 1:
            break
    resultjuge(minitbookaddr__, signal, factornumber1, rownumber_, max_number, listtemple, get_num_lista21)


def extractdata(minitbookaddr__, signal, factornumber1, rownumber_, mproject1=None, get_num_lista21=None):
    if mproject1 is None:
        mproject1 = win32com.client.Dispatch('{D3407EDD-1BAE-4E00-A512-2673EBF849CE}').ActiveProject
    if get_num_lista21 is None:
        get_num_lista21 = [0 for _ in range(rownumber_)]
    max_number = numpy.zeros(factornumber1)
    listtemple = []
    for j1 in range(1, mproject1.Commands.Count + 1):
        command = mproject1.Commands(j1)
        signals = 0
        for e in range(1, command.Outputs.Count + 1):
            number1 = command.Outputs(e).OutputType
            if number1 == 1:
                responsestr = str(command.Outputs(e).Table.Text)
                if 'Response Table for Means' in responsestr:
                    loggenera('均值相应表：' + '\n' + responsestr)
                    maxindexarray = maxindex(factornumber1, responsestr, max_number)
                    listtemple = writemaxindex(minitbookaddr__, rownumber_, factornumber1, maxindexarray,
                                               get_num_lista21)
                    signals = 1
                    break
        if signals == 1:
            break
    resultjuge(minitbookaddr__, signal, factornumber1, rownumber_, max_number, listtemple, get_num_lista21)


def extractdatasinglesmalnum(minitbookaddr__, signal, factornumber1, rownumber_, mproject1=None, get_num_lista21=None):
    if mproject1 is None:
        mproject1 = win32com.client.Dispatch('{D3407EDD-1BAE-4E00-A512-2673EBF849CE}').ActiveProject
    if get_num_lista21 is None:
        get_num_lista21 = [0 for _ in range(rownumber_)]
    max_number = numpy.zeros(factornumber1)
    listtemple = []
    for j1 in range(1, mproject1.Commands.Count + 1):
        command = mproject1.Commands(j1)
        signals = 0
        for e in range(1, command.Outputs.Count + 1):
            number1 = command.Outputs(e).OutputType
            if number1 == 1:
                responsestr = str(command.Outputs(e).Table.Text)
                if 'Response Table for Means' in responsestr:
                    loggenera('均值相应表：' + '\n' + responsestr)
                    maxindexarray = maxindex(factornumber1, responsestr, max_number)
                    listtemple = writemaxindex(minitbookaddr__, rownumber_, factornumber1, maxindexarray,
                                               get_num_lista21)
                    signals = 1
                    break
        if signals == 1:
            break
    resultjugesingle(minitbookaddr__, signal, factornumber1, rownumber_, max_number, listtemple, get_num_lista21)


def extractdatasinglebignu(minitbookaddr__, signal, factornumber1, rownumber_, mproject1=None, get_num_lista21=None):
    if mproject1 is None:
        mproject1 = win32com.client.Dispatch('{D3407EDD-1BAE-4E00-A512-2673EBF849CE}').ActiveProject
    if get_num_lista21 is None:
        get_num_lista21 = [0 for _ in range(rownumber_)]
    max_number = numpy.zeros(factornumber1)
    listtemple = []
    for j1 in range(1, mproject1.Commands.Count + 1):
        command = mproject1.Commands(j1)
        signals = 0
        for e in range(1, command.Outputs.Count + 1):
            number1 = command.Outputs(e).OutputType
            if number1 == 1:
                responsestr = str(command.Outputs(e).Table.Text)
                if 'Response Table for Means' in responsestr:
                    loggenera('均值相应表：' + '\n' + responsestr)
                    maxindexarray = maxindextwolev(factornumber1, responsestr, max_number)
                    listtemple = writemaxindex(minitbookaddr__, rownumber_, factornumber1, maxindexarray,
                                               get_num_lista21)
                    signals = 1
                    break
        if signals == 1:
            break
    resultjugesingle(minitbookaddr__, signal, factornumber1, rownumber_, max_number, listtemple, get_num_lista21)


def resultjugesingle(minitbookaddr__, signal, factornumber1, rownumber_, max_number=None, listtemple=None,
                     get_num_lista21=None):
    if max_number is None:
        max_number = numpy.zeros(factornumber1)
    if listtemple is None:
        listtemple = [0 * i1 for i1 in range(rownumber_)]
    if get_num_lista21 is None:
        get_num_lista21 = [0 * i1 for i1 in range(rownumber_)]
    logdebug(listtemple)
    curmaxv = round(max(max_number), 0)
    maxvalue = 0
    previousv = 0
    if signal == 0:
        maxvalue = round(getdata(minitbookaddr__, 20, 9), 0)
        previousv = round(getdata(minitbookaddr__, 21, 9), 0)
        repicheck = (previousv == curmaxv) and (listtemple == get_num_lista21)
        setdata(minitbookaddr__, 21, 9, curmaxv)

        if curmaxv > maxvalue:
            setdata(minitbookaddr__, 20, 9, curmaxv)
        if repicheck or curmaxv <= previousv:
            mutation(minitbookaddr__, signal, rownumber_, listtemple)
            setdata(minitbookaddr__, 16, 9, 1)
    writelist(minitbookaddr__, rownumber_, listtemple, get_num_lista21)
    loggenera('最大值：' + str(maxvalue) + ' ' + '前次计算值：' + str(previousv) + ' ' + '当前计算值：' + str(curmaxv))


def resultjuge(minitbookaddr__, signal, factornumber1, rownumber_, max_number=None, listtemple=None,
               get_num_lista21=None):
    if max_number is None:
        max_number = numpy.zeros(factornumber1)
    if listtemple is None:
        listtemple = [0 * i1 for i1 in range(rownumber_)]
    if get_num_lista21 is None:
        get_num_lista21 = [0 * i1 for i1 in range(rownumber_)]
    curmaxv = round(max(max_number), 0)
    maxvalue = 0
    previousv = 0
    if signal == 0:
        maxvalue = round(getdata(minitbookaddr__, 20, 9), 0)
        previousv = round(getdata(minitbookaddr__, 21, 9), 0)
        repicheck = (previousv == curmaxv) and (listtemple == get_num_lista21)
        setdata(minitbookaddr__, 21, 9, curmaxv)

        if curmaxv > maxvalue:
            setdata(minitbookaddr__, 20, 9, curmaxv)
        if repicheck or curmaxv <= previousv:
            mutation(minitbookaddr__, signal, rownumber_, listtemple)
            setdata(minitbookaddr__, 16, 9, 1)

    elif signal == 1:
        maxvalue = round(getdata(minitbookaddr__, 18, 9), 0)
        previousv = round(getdata(minitbookaddr__, 19, 9), 0)
        repicheck = (previousv == curmaxv) and (listtemple == get_num_lista21)
        setdata(minitbookaddr__, 19, 9, curmaxv)

        if curmaxv > maxvalue:
            setdata(minitbookaddr__, 18, 9, curmaxv)
        if curmaxv > maxvalue * 0.97:
            setdata(minitbookaddr__, 17, 9, 1)
        elif repicheck or curmaxv <= previousv:
            mutation(minitbookaddr__, signal, rownumber_, listtemple)
    writelist(minitbookaddr__, rownumber_, listtemple, get_num_lista21)

    loggenera('最大值：' + str(maxvalue) + ' ' + '前次计算值：' + str(previousv) + ' ' + '当前计算值：' + str(curmaxv))


def mutation(minitbookaddr_, signal, rownumber_, listtemple=None):
    if listtemple is None:
        listtemple = [0 * i1 for i1 in range(rownumber_)]
    steplist = []
    extract_data_from_excellist(minitbookaddr_, 2, 1, 8, 9, 'Sheet', steplist)
    quickstep = steplist[0]
    transtep = steplist[1]
    coefficient = mutatchange(signal, listtemple[0], quickstep, transtep, minitbookaddr_)
    for ii in range(rownumber_):
        random.randint(0, 1)
        if random.randint(0, 1) == 0:
            randomratio = 1 - coefficient
        else:
            randomratio = 1 + coefficient
        listtemple[ii] = listtemple[ii] * randomratio


def coefficientratiochange(signal, turnnum, quickstep, transtep, minitbookaddr_):
    coefficientcal = 0.01
    quickstep1 = quickstep
    transtep1 = transtep
    if quickstep1 == 0:
        quickstep1 = 5
    if transtep1 == 0:
        transtep1 = 1
    if turnnum > 0:
        if signal == 0:
            maxdata = getdata(minitbookaddr_, 21, 9)
            if maxdata is None or maxdata <= 0:
                ratiochange = 200
            else:
                ratiochange = 700 / maxdata
            if ratiochange > 100:
                coefficientcal = 0.5
            else:
                if 0 < transtep1 < 1:
                    coefficientcal = transtep1 * ratiochange
                else:
                    transtepchange = transtep1 * ratiochange
                    while turnnum * coefficientcal < transtepchange:
                        coefficientcal = coefficientcal + 0.01

        if signal == 1:
            maxdata = getdata(minitbookaddr_, 19, 9)
            if maxdata is None or maxdata <= 0:
                ratiochange = 1000
            else:
                ratiochange = 700 / maxdata
            if ratiochange > 100:
                coefficientcal = 0.5
            else:
                if 0 < quickstep1 < 1:
                    coefficientcal = quickstep1 * ratiochange
                else:
                    quickstepchange = quickstep1 * ratiochange
                    while turnnum * coefficientcal < quickstepchange:
                        coefficientcal = coefficientcal + 0.01

    if coefficientcal > 0.5:
        coefficientcal = 0.5
    if coefficientcal < 0.01:
        coefficientcal = 0.01
    return coefficientcal


def mutatchange(signal, turnnum, quickstep, transtep, minitbookaddr_):
    mutatchara = 0.5
    quickstep1 = quickstep
    transtep1 = transtep

    if quickstep1 == 0:
        quickstep1 = 0.5
    if transtep1 == 0:
        transtep1 = 0.5
    if turnnum > 0:
        if signal == 0:
            maxdata = getdata(minitbookaddr_, 21, 9)
            if maxdata is None:
                maxdata = 0
            if maxdata < 450:
                ratiochange = 1.5
            elif maxdata < 720:
                ratiochange = 1.2
            else:
                ratiochange = 1
            mutatchara = transtep1 * ratiochange

        if signal == 1:
            maxdata = getdata(minitbookaddr_, 19, 9)
            if maxdata is None:
                maxdata = 0
            if maxdata < 450:
                ratiochange = 1.5
            elif maxdata < 720:
                ratiochange = 1.2
            else:
                ratiochange = 1
            mutatchara = quickstep1 * ratiochange
    if mutatchara >= 1:
        mutatchara = 0.99
    elif mutatchara <= 0:
        mutatchara = 0.01

    return mutatchara


def writelist(minitbookaddr_, rownumber_, listtemple=None, get_num_lista21=None):
    if listtemple is None:
        listtemple = [0 * i1 for i1 in range(rownumber_)]
    if get_num_lista21 is None:
        get_num_lista21 = [0 * i1 for i1 in range(rownumber_)]
    while 1:
        try:
            wb = load_workbook(minitbookaddr_)
            for i1 in range(rownumber_):
                if get_num_lista21[i1] != 0:
                    while 1:
                        wb['Sheet'].cell(row=i1 + 2, column=2).value = listtemple[i1]
                        if wb['Sheet'].cell(row=i1 + 2, column=2).value == listtemple[i1]:
                            break
            wb.save(minitbookaddr_)
            logdebug(get_num_lista21)
            loggenera(listtemple)
            time.sleep(0.1)
            wb.close()
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(3)
            continue


def writemaxlist(minitbookaddr_, rowsele, factornumber1, max_number=None):
    if max_number is None:
        max_number = numpy.zeros(factornumber1)
    while 1:
        try:
            wb = load_workbook(minitbookaddr_)
            for i1 in range(factornumber1):
                wb['Sheet'].cell(row=rowsele, column=i1 + 1).value = max_number[i1]
            wb.save(minitbookaddr_)
            time.sleep(0.1)
            wb.close()
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(3)
            continue


def readdatainput(minitbookaddr_, rownumber_, get_num_lista=None, num_list_generate=None):
    if get_num_lista is None:
        get_num_lista = [0 * i for i in range(rownumber_)]
    if num_list_generate is None:
        num_list_generate = numpy.zeros((rownumber_, 3))

    while 1:
        try:
            wb = load_workbook(minitbookaddr_)
            for i2 in range(rownumber_):
                wb['Sheet'].cell(row=2 + i2, column=2).value = get_num_lista[i2]
                wb['Sheet'].cell(row=2 + i2, column=4).value = num_list_generate[i2][0]
                wb['Sheet'].cell(row=2 + i2, column=5).value = num_list_generate[i2][1]
                wb['Sheet'].cell(row=2 + i2, column=6).value = num_list_generate[i2][2]
            wb.save(minitbookaddr_)
            time.sleep(0.1)
            wb.close()
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(3)
            continue


def creatminitexcel(workbookaddr_, namestrin, namelist_=None):
    if namelist_ is None:
        namelist_ = []
    wb = Workbook()
    minitabnamlistin = ['性能曲线描点个数', '额定转矩mN.m', '允许齿槽转矩mN.m', '单相等待时间调整系数（默认5）', '快算步长（1-最大匝数）', '细算步长（1-最大匝数）',
                        '快算变异百分比（0.1-0.99）', '细算变异百分比（0.1-0.99）', '三相等待时间调整系数（默认1）', '优化次数（1-500）', '计算次数', '快算计次',
                        '细算计次', '绿色：输入区域', '浅蓝色：说明区']
    minitabnamlistinadd = ['下方输入测试时马达匝数']
    minitabnamlist = ['序号', '中值', r'因素\水平', '1', '2', '3', '下限', '上限', '数值', '名称']

    sheetnam = 'Sheet'
    wb.create_sheet(sheetnam, 1)
    minitabnlen = len(minitabnamlist)
    minitabninlen = len(minitabnamlistin)
    seclist = [i + 1 for i in range(namelist_.index('0'))]
    thin = Side(border_style='thin', color='000000')
    double = Side(border_style='medium', color='000000')
    borderdim = [thin, double]
    for i in range(namelist_.index('0') + 1):
        for j in range(10):
            row1 = i + 1
            column1 = j + 1
            cell = wb[sheetnam].cell(row=row1, column=column1)
            if i % 2 == 0:
                cell.border = Border(top=borderdim[1], left=borderdim[0], right=borderdim[0])
            else:
                cell.border = Border(top=borderdim[0], left=borderdim[0], right=borderdim[0])

    fillc = PatternFill('solid', fgColor='cffcfc')
    fillcinp = PatternFill('solid', fgColor='71f495')

    for j in range(minitabnlen):
        row1 = 1
        column1 = j + 1
        cell = wb[sheetnam].cell(row=row1, column=column1)
        cell.value = minitabnamlist[j]
        cell.fill = fillc
    for j in range(minitabninlen):
        row1 = j + 2
        column1 = 10
        cell = wb[sheetnam].cell(row=row1, column=column1)
        cell.value = minitabnamlistin[j]
        cell.fill = fillc

    for j in range(namelist_.index('0')):
        row1 = j + 2
        column1 = 3
        cell = wb[sheetnam].cell(row=row1, column=column1)
        cell.value = namelist_[j]
        cell.fill = fillc

    for j in range(namelist_.index('0')):
        row1 = j + 2
        column1 = 1
        cell = wb[sheetnam].cell(row=row1, column=column1)
        cell.value = seclist[j]
        cell.fill = fillc
    for j in range(namelist_.index('0')):
        row1 = j + 2
        column1 = 2
        cell = wb[sheetnam].cell(row=row1, column=column1)
        cell.value = 0
        cell.fill = fillcinp

    cell1 = wb[sheetnam].cell(row=23, column=10)
    cell1.value = minitabnamlistinadd[0]
    cell1.fill = fillc

    cell2 = wb[sheetnam].cell(row=24, column=10)
    cell2.value = 0
    cell2.fill = fillcinp

    for j in range(namelist_.index('0')):
        for i in range(2):
            row1 = j + 2
            column1 = i + 7
            cell = wb[sheetnam].cell(row=row1, column=column1)
            cell.value = 0
            cell.fill = fillcinp

    for j in range(10):
        row1 = j + 2
        column1 = 9
        cell = wb[sheetnam].cell(row=row1, column=column1)
        cell.fill = fillcinp

    for j in range(14):
        row1 = j + 12
        column1 = 9
        cell = wb[sheetnam].cell(row=row1, column=column1)
        cell.value = 0
        cell.fill = fillc

    ws = wb[sheetnam]
    ws.column_dimensions['C'].width = 16
    wb.save(workbookaddr_)

    dv = DataValidation(type='list', formula1=namestrin, prompt='选择模板', promptTitle='下拉列表')
    modecell = ws['J17']
    modecell.fill = fillcinp
    dv.add(modecell)
    ws.add_data_validation(dv)

    dv2 = DataValidation(type='list', formula1='"1,2,3"', prompt='1：计算仅改变绕线时，最大开环匝数; \n2：计算仅改变绕线时，最大背压匝数; \n3：计算仅改变绕线时，最大开环匝数和最大背压匝数', promptTitle='下拉列表')
    modecell2 = ws['J21']
    modecell2.fill = fillcinp
    dv2.add(modecell2)
    ws.add_data_validation(dv2)

    ws.views.sheetView[0].zoomScale = 120
    wb.save(workbookaddr_)


def testresultexcelhandle(testresultaddr_, testresultafteraddr_, number):
    while 1:
        try:
            xlapp = win32com.client.Dispatch("Excel.Application")
            xlapp.Visible = False
            wb = xlapp.Workbooks.Open(testresultaddr_)
            overflowsign = 0
            if (wb.Worksheets('sheet1').Cells(17, 29).Value * wb.Worksheets('sheet1').Cells(17, 6).Value / (wb.Worksheets('sheet1').Cells(17, 9).Value * 973.8135)) > 100:
                overflowsign = 1
            dataarray = numpy.zeros((number, 9, 7))

            for i in range(number):
                sheetn = 'Sheet%d' % (i*2 + 1)
                if wb.Worksheets(sheetn).Cells(17, 29).Value == 0 or wb.Worksheets(sheetn).Cells(17, 29).Value is None:
                    break
                wb.Worksheets(sheetn).Cells(15, 30).Value = 'Mechanic Power (w)'
                wb.Worksheets(sheetn).Cells(15, 31).Value = 'Fan flow Power (w)'
                wb.Worksheets(sheetn).Cells(15, 32).Value = 'Motor Eff. (%)'
                wb.Worksheets(sheetn).Cells(15, 33).Value = 'Fan Eff. (%)'

                for e in range(9):

                    row1 = 17 + e
                    if overflowsign == 1:
                        tempvalue = wb.Worksheets(sheetn).Cells(row1, 29).Value/10
                        wb.Worksheets(sheetn).Cells(row1, 29).Value = tempvalue
                    wb.Worksheets(sheetn).Cells(row1, 30).Value = wb.Worksheets(sheetn).Cells(row1, 6).Value * wb.Worksheets(sheetn).Cells(row1, 29).Value / 97381.35
                    wb.Worksheets(sheetn).Cells(row1, 31).Value = wb.Worksheets(sheetn).Cells(row1, 4).Value * wb.Worksheets(sheetn).Cells(row1, 5).Value * 0.00449
                    wb.Worksheets(sheetn).Cells(row1, 32).Value = wb.Worksheets(sheetn).Cells(row1, 30).Value * 100 / wb.Worksheets(sheetn).Cells(row1, 9).Value
                    wb.Worksheets(sheetn).Cells(row1, 33).Value = wb.Worksheets(sheetn).Cells(row1, 31).Value * 100 / wb.Worksheets(sheetn).Cells(row1, 30).Value
                    dataarray[i][e][0] = wb.Worksheets(sheetn).Cells(row1, 6).Value
                    dataarray[i][e][1] = wb.Worksheets(sheetn).Cells(row1, 8).Value
                    dataarray[i][e][2] = round(wb.Worksheets(sheetn).Cells(row1, 29).Value / 10.197, 2)
                    dataarray[i][e][3] = wb.Worksheets(sheetn).Cells(row1, 14).Value
                    dataarray[i][e][4] = wb.Worksheets(sheetn).Cells(row1, 32).Value
                    dataarray[i][e][5] = wb.Worksheets(sheetn).Cells(row1, 4).Value
                    dataarray[i][e][6] = wb.Worksheets(sheetn).Cells(row1, 5).Value



            wb.SaveAs(testresultafteraddr_)
            time.sleep(0.1)
            wb.Close()
            xlapp.Quit()
            return dataarray
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            logdebug('  Exception: ' + str(Exception))
            time.sleep(3)
            continue


def inputdatahandle(minitbookaddr_, maxsspeed, maxscurrent, maxstorque, phaseinpu):
    while 1:
        try:
            wb111 = load_workbook(minitbookaddr_)
            if phaseinpu != 3 and phaseinpu != 4:
                turn = round(wb111['Sheet'].cell(24, 10).value, 0)
                wb111['Sheet'].cell(2, 7).value = turn
                wb111['Sheet'].cell(2, 8).value = turn
                wb111['Sheet'].cell(2, 2).value = turn

                wb111['Sheet'].cell(3, 7).value = maxsspeed
                wb111['Sheet'].cell(3, 8).value = maxsspeed
                wb111['Sheet'].cell(3, 2).value = maxsspeed

                advangle = round(wb111['Sheet'].cell(11, 7).value, 0)
                wb111['Sheet'].cell(11, 7).value = advangle
                wb111['Sheet'].cell(11, 8).value = advangle
                wb111['Sheet'].cell(11, 2).value = advangle

                wb111['Sheet'].cell(13, 7).value = maxscurrent
                wb111['Sheet'].cell(13, 8).value = maxscurrent
                wb111['Sheet'].cell(13, 2).value = maxscurrent

                wb111['Sheet'].cell(3, 9).value = maxstorque/10.197
                wb111['Sheet'].cell(4, 9).value = maxstorque

                wb111.save(minitbookaddr_)
                time.sleep(0.1)
                if wb111['Sheet'].cell(2, 2).value == turn:
                    break
                else:
                    loggenera('excel write failed')

            wb111.close()
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def inputdatahandlehigeff(minitbookaddr_, turns, maxsspeed, maxscurrent, maxstorque, phaseinpu):
    while 1:
        try:
            wb111 = load_workbook(minitbookaddr_)
            if phaseinpu != 3 and phaseinpu != 4:
                wb111['Sheet'].cell(2, 7).value = round(turns*0.6, 0)
                wb111['Sheet'].cell(2, 8).value = turns
                wb111['Sheet'].cell(2, 2).value = turns

                wb111['Sheet'].cell(3, 7).value = maxsspeed
                wb111['Sheet'].cell(3, 8).value = maxsspeed
                wb111['Sheet'].cell(3, 2).value = maxsspeed

                advangle = 0.1
                wb111['Sheet'].cell(11, 7).value = 0.1
                wb111['Sheet'].cell(11, 8).value = 15
                wb111['Sheet'].cell(11, 2).value = 1

                wb111['Sheet'].cell(13, 7).value = maxscurrent*1.2
                wb111['Sheet'].cell(13, 8).value = maxscurrent*1.8
                wb111['Sheet'].cell(13, 2).value = maxscurrent

                wb111['Sheet'].cell(3, 9).value = maxstorque/10.197
                wb111['Sheet'].cell(4, 9).value = maxstorque

                wb111.save(minitbookaddr_)
                time.sleep(0.1)
                if wb111['Sheet'].cell(2, 8).value == turns:
                    break
                else:
                    loggenera('excel write failed')

            wb111.close()
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def creatplotexcel(plotorkbookaddr_, calcunum, number, rowstart1, rowstart2, rowstart3, returnrown_, toque, cogging,
                   namelist_=None):
    if namelist_ is None:
        namelist_ = []
    wb = Workbook()
    namlen = len(namelist_)
    for i in range(number):
        sheetn = 'Sheet%d' % (i + 1)
        wb.create_sheet(sheetn, i + 1)
        for j in range(namlen):
            row1 = 1
            column1 = j + 1
            wb[sheetn].cell(row=row1, column=column1).value = namelist_[j]
    wb.save(plotorkbookaddr_)

    efficstr = ['效率-%s曲线%s', '转矩-%s曲线%s', '齿槽转矩-%s曲线%s']
    limitlist = [[60, 100], [toque * 0.5, toque * 1.5], [cogging * 0.5, cogging * 1.5]]
    chartplacestr = ['A', 'J', 'S']
    columx = namelist_.index('母线平均电流') + 1
    columx1 = len(namelist_) + 2
    columy = [namelist_.index('效率') + 1, namelist_.index('转矩') + 1, namelist_.index('齿槽转矩波动') + 1]
    chartplacestr1 = [j + str(1) for j in chartplacestr]
    efficstr1 = [i % ('电流', '(田口优化)') for i in efficstr]
    for i in range(3):
        scatterchartcreat(plotorkbookaddr_, chartplacestr1[i], efficstr1[i], number, rowstart1, calcunum, columx,
                          columy[i], limitlist[i], wb)

    chartplacestr2 = [j + str(18) for j in chartplacestr]
    efficstr2 = [i % ('序号', '(细算库存最优)') for i in efficstr]
    for i in range(3):
        scatterchartcreat(plotorkbookaddr_, chartplacestr2[i], efficstr2[i], 1, rowstart2, returnrown_, columx1,
                          columy[i], limitlist[i], wb)

    chartplacestr3 = [j + str(35) for j in chartplacestr]
    efficstr3 = [i % ('序号', '(粗算库存最优)') for i in efficstr]
    for i in range(3):
        scatterchartcreat(plotorkbookaddr_, chartplacestr3[i], efficstr3[i], 1, rowstart3, returnrown_, columx1,
                          columy[i], limitlist[i], wb)


def scatterchartcreat(plotorkbookaddr_, chartplacestr, efficstr, number, rowstart, calcunum, columx, columy,
                      limitlist=None, wb=None):
    if wb is None:
        wb = Workbook()
    if limitlist is None:
        limitlist = [0, 0]
    effchart = ScatterChart(scatterStyle='line')
    effchart.title = efficstr
    effchart.style = 13
    effchart.y_axis.scaling.min = limitlist[0]
    effchart.y_axis.scaling.max = limitlist[1]
    lineprop = LineProperties(w=1, prstDash="solid")
    for i in range(number):
        sheetn = 'Sheet%d' % (i + 1)
        ws = wb[sheetn]
        xvalues = Reference(ws, min_row=rowstart, max_row=calcunum + rowstart, min_col=columx, max_col=columx)
        yvalues = Reference(ws, min_row=rowstart, max_row=calcunum + rowstart, min_col=columy, max_col=columy)
        serie = Series(yvalues, xvalues, title=str(i + 1))
        serie.graphicalProperties.line = lineprop
        effchart.series.append(serie)

    ws1 = wb['Sheet']
    ws1.add_chart(effchart, chartplacestr)
    ws1.views.sheetView[0].zoomScale = 120
    wb.save(plotorkbookaddr_)


def plotcurve(workbookaddr_, plotorkbookaddr_, index, rownumber1, rownumber_):
    sheet = 'Sheet' + index
    while 1:
        try:
            wb3 = load_workbook(workbookaddr_)
            wbpl = load_workbook(plotorkbookaddr_)
            for i2 in range(rownumber1):
                for j1 in range(rownumber_ + 12):
                    row1 = i2 + 6
                    column1 = j1 + 2
                    row2 = i2 + 2
                    column2 = j1 + 1
                    if wb3['Sheet'].cell(row=row1, column=column1).value is None:
                        wbpl[sheet].cell(row=row2, column=column2).value = 0
                    else:
                        wbpl[sheet].cell(row=row2, column=column2).value = wb3['Sheet'].cell(row=row1,
                                                                                             column=column1).value
            wbpl.save(plotorkbookaddr_)
            time.sleep(0.1)
            wbpl.close()
            wb3.close()
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def exchangeexcel(minitbookaddr_, minitbookaddr_temp, rowstar, rownumberend, colstar, columnnumberend):
    while 1:
        try:
            wb3 = load_workbook(minitbookaddr_)
            wbpl = load_workbook(minitbookaddr_temp)
            for i2 in range(rowstar, rownumberend + 1):
                for j1 in range(colstar, columnnumberend + 1):
                    row1 = i2
                    column1 = j1
                    wbpl['Sheet'].cell(row=row1, column=column1).value = wb3['Sheet'].cell(row=row1,
                                                                                           column=column1).value
            wbpl.save(minitbookaddr_temp)
            time.sleep(0.1)
            wbpl.close()
            wb3.close()
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def setzero(minitbookaddr_):
    while 1:
        try:
            wb111 = load_workbook(minitbookaddr_)
            wb111['Sheet'].cell(row=22, column=9).value = 0
            wb111['Sheet'].cell(row=23, column=9).value = 0
            wb111['Sheet'].cell(row=24, column=9).value = 0
            wb111['Sheet'].cell(row=25, column=9).value = 0
            wb111.save(minitbookaddr_)
            time.sleep(0.1)
            wb111.close()
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def appendlist(minitbookaddr_):
    listsin11 = []
    while 1:
        try:
            wb111 = load_workbook(minitbookaddr_)
            listsin11.append(wb111['Sheet'].cell(row=22, column=9).value)
            listsin11.append(wb111['Sheet'].cell(row=23, column=9).value)
            listsin11.append(wb111['Sheet'].cell(row=24, column=9).value)
            listsin11.append(wb111['Sheet'].cell(row=25, column=9).value)
            time.sleep(0.1)
            wb111.close()
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(3)
            continue
    return listsin11


def setdata(minitbookaddr_, row, column, times1):
    while 1:
        try:
            wb111 = load_workbook(minitbookaddr_)
            wb111['Sheet'].cell(row, column).value = times1
            wb111.save(minitbookaddr_)
            time.sleep(0.1)
            if wb111['Sheet'].cell(row, column).value == times1:
                break
            else:
                loggenera('excel write failed')
            wb111.close()
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def getdata(minitbookaddr_, row, column):
    while 1:
        try:
            wb111 = load_workbook(minitbookaddr_)
            times1 = wb111['Sheet'].cell(row, column).value
            time.sleep(0.1)
            if times1 == wb111['Sheet'].cell(row, column).value:
                wb111.close()
                return times1
            else:
                loggenera('excel read failed')
                wb111.close()
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def write_data_to_excel_after(rownumber_, workbookaddr_, rownumber1, rownumberall, columnnumber21,
                              return_num_list1=None):
    if return_num_list1 is None:
        return_num_list1 = numpy.zeros((rownumberall, columnnumber21))
    while 1:
        try:
            wb11 = load_workbook(workbookaddr_)
            zerocheck = 0
            for j in range(columnnumber21):
                row1 = rownumber1 + 6
                column1 = j + rownumber_ + 3
                wb11['Sheet'].cell(row=row1, column=column1).value = return_num_list1[rownumber1][j]
                j += 1
                zerocheck += return_num_list1[rownumber1][j]
            wb11.save(workbookaddr_)
            time.sleep(0.1)
            wb11.close()
            return zerocheck
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def write_data_to_excel(rownumberexcel, rownumber1, rownumberall, columnnumber11, workbookaddr11, num_list=None,
                        wb11=None):
    if num_list is None:
        num_list = numpy.zeros((rownumberall, columnnumber11))
    if wb11 is None:
        wb11 = Workbook()
    while 1:
        try:
            for j in range(columnnumber11):
                row1 = rownumberexcel + 6
                column1 = j + 2
                wb11['Sheet'].cell(row=row1, column=column1).value = num_list[rownumber1][j]
            wb11.save(workbookaddr11)
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def write_data_to_excelholexc(rownumberexcel, rownumber1, rownumberall, columnnumber11, rowoffset, coloffset, sheet,
                              workbookaddr11, num_list=None):
    if num_list is None:
        num_list = numpy.zeros((rownumberall, columnnumber11))
    while 1:
        try:
            wb11 = load_workbook(workbookaddr11)
            for j in range(columnnumber11):
                row1 = rownumberexcel + rowoffset
                column1 = j + coloffset
                wb11[sheet].cell(row=row1, column=column1).value = num_list[rownumber1][j]
            wb11.save(workbookaddr11)
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def write_data_to_excelhol(minitbookaddr_, rownumber0, columnnumber0, rowoffset, coloffset, sheet, num_array=None):
    if num_array is None:
        num_array = numpy.zeros((rownumber0, columnnumber0))
    while 1:
        try:
            wb = load_workbook(minitbookaddr_)
            for i in range(rownumber0):
                for j in range(columnnumber0):
                    row1 = i + rowoffset
                    column1 = j + coloffset
                    wb[sheet].cell(row=row1, column=column1).value = num_array[i][j]
            wb.save(minitbookaddr_)
            time.sleep(0.1)
            wb.close()
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def extract_data_from_excelarray(minitbookaddr_, rownumber0, columnnumber0, rowoffset, coloffset, sheet,
                                 num_array=None):
    # save Sheet to num_list1
    if num_array is None:
        num_array = numpy.zeros((rownumber0, columnnumber0))

    while 1:
        try:
            wb = load_workbook(minitbookaddr_)
            for i in range(rownumber0):
                for j in range(columnnumber0):
                    row1 = i + rowoffset
                    column1 = j + coloffset
                    if wb[sheet].cell(row=row1, column=column1).value is None:
                        num_array[i][j] = 0
                    else:
                        num_array[i][j] = wb[sheet].cell(row=row1, column=column1).value
            time.sleep(0.1)
            wb.close()
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def extract_data_from_excellist(minitbookaddr_, rownumber0, columnnumber0, rowoffset, coloffset, sheet, num_list0=None):
    # save Sheet to num_list1
    if num_list0 is None:
        num_list0 = []
    num_list0.clear()
    while 1:
        try:
            wb11 = load_workbook(minitbookaddr_)
            for i1 in range(rownumber0):
                for j1 in range(columnnumber0):
                    row1 = i1 + rowoffset
                    column1 = j1 + coloffset
                    if wb11[sheet].cell(row=row1, column=column1).value is None:
                        num_list0.append(0)
                    else:
                        num_list0.append(wb11[sheet].cell(row=row1, column=column1).value)
            time.sleep(0.1)
            wb11.close()
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def just_open(filenamem):
    while 1:
        try:
            xlapp = win32com.client.Dispatch("Excel.Application")
            xlapp.Visible = True
            xlbook = xlapp.Workbooks.Open(filenamem)
            time.sleep(0.1)
            xlbook.Save()
            time.sleep(0.01)
            xlbook.Close()
            break
        except (PermissionError, BadZipFile, EOFError, com_error, zlib.error):
            time.sleep(1)
            continue


def extractdatatest(mproject=None):
    if mproject is None:
        mproject = win32com.client.Dispatch('{D3407EDD-1BAE-4E00-A512-2673EBF849CE}').ActiveProject

    for j1 in range(1, mproject.Commands.Count + 1):
        command = mproject.Commands(j1)
        for i1 in range(1, command.Outputs.Count + 1):
            number1 = command.Outputs(i1).OutputType
            if number1 == 1:
                responsestr = str(command.Outputs(i1).Table.Text)
                loggenera(j1)
                loggenera(i1)
                loggenera(command.Tag)
                loggenera(command.Name)
                loggenera(responsestr)


def hhcommand(rownumber_, factornumber1, mproject=None):
    if mproject is None:
        mproject = win32com.client.Dispatch('{D3407EDD-1BAE-4E00-A512-2673EBF849CE}').ActiveProject

    mproject.ExecuteCommand(analyzecda(rownumber_ + 5, 1, factornumber1) + semicolon + space + space +
                            designcda(1, factornumber1) + semicolon + space + space +
                            'Snratio C67;' + '  Means C68;' + space + space +
                            'Higher' + semicolon + space + space +
                            plotmapnt)
    mproject.ExecuteCommand(analyzecda(rownumber_ + 4, 1, factornumber1) + semicolon + space + space +
                            designcda(1, factornumber1) + semicolon + space + space +
                            'Snratio C69;' + '  Means C70;' + space + space +
                            'Higher' + semicolon + space + space +
                            plotmap)
    mproject.ExecuteCommand('Erase C400 C399;' + 'Abort.')


def lhcommand(rownumber_, factornumber1, mproject=None):
    if mproject is None:
        mproject = win32com.client.Dispatch('{D3407EDD-1BAE-4E00-A512-2673EBF849CE}').ActiveProject

    mproject.ExecuteCommand(analyzecda(rownumber_ + 5, 1, factornumber1) + semicolon + space + space +
                            designcda(1, factornumber1) + semicolon + space + space +
                            'Snratio C67;' + '  Means C68;' + space + space +
                            'Lower' + semicolon + space + space +
                            plotmapnt)
    mproject.ExecuteCommand(analyzecda(rownumber_ + 4, 1, factornumber1) + semicolon + space + space +
                            designcda(1, factornumber1) + semicolon + space + space +
                            'Snratio C67;' + '  Means C68;' + space + space +
                            'Higher' + semicolon + space + space +
                            plotmap)
    mproject.ExecuteCommand('Erase C400 C399;' + 'Abort.')


def hhcommandb(rownumber_, factornumber1, mproject=None):
    if mproject is None:
        mproject = win32com.client.Dispatch('{D3407EDD-1BAE-4E00-A512-2673EBF849CE}').ActiveProject

    mproject.ExecuteCommand(analyzecda(rownumber_ + 5, 2, factornumber1 + 1) + semicolon + space + space +
                            designcda(2, factornumber1 + 1) + semicolon + space + space +
                            'Snratio C67;' + '  Means C68;' + space + space +
                            'Higher' + semicolon + space + space +
                            plotmapnt)
    mproject.ExecuteCommand(analyzecda(rownumber_ + 4, 2, factornumber1 + 1) + semicolon + space + space +
                            designcda(2, factornumber1 + 1) + semicolon + space + space +
                            'Snratio C69;' + '  Means C70;' + space + space +
                            'Higher' + semicolon + space + space +
                            plotmap)
    mproject.ExecuteCommand('Erase C400 C399;' + 'Abort.')


def lhcommandb(rownumber_, factornumber1, mproject=None):
    if mproject is None:
        mproject = win32com.client.Dispatch('{D3407EDD-1BAE-4E00-A512-2673EBF849CE}').ActiveProject

    mproject.ExecuteCommand(analyzecda(rownumber_ + 5, 2, factornumber1 + 1) + semicolon + space + space +
                            designcda(2, factornumber1 + 1) + semicolon + space + space +
                            'Snratio C67;' + '  Means C68;' + space + space +
                            'Lower' + semicolon + space + space +
                            plotmapnt)
    mproject.ExecuteCommand(analyzecda(rownumber_ + 4, 2, factornumber1 + 1) + semicolon + space + space +
                            designcda(2, factornumber1 + 1) + semicolon + space + space +
                            'Snratio C67;' + '  Means C68;' + space + space +
                            'Higher' + semicolon + space + space +
                            plotmap)
    mproject.ExecuteCommand('Erase C400 C399;' + 'Abort.')


def minitabop(minitabaddr1):
    while 1:
        try:
            minitab = win32com.client.Dispatch('{D3407EDD-1BAE-4E00-A512-2673EBF849CE}')
            time.sleep(1)
            minitab.Open(minitabaddr1)
            time.sleep(1)
            minitab.UserInterface.Visible = True
            time.sleep(1)
            mproject = minitab.ActiveProject
            return minitab, mproject
        except (PermissionError, BadZipFile, EOFError, com_error, AttributeError):
            time.sleep(1)
            continue


def minitabopgen(minitabaddr_):
    while 1:
        try:
            minitab = win32com.client.Dispatch('{D3407EDD-1BAE-4E00-A512-2673EBF849CE}')
            time.sleep(1)
            minitab.New()
            time.sleep(1)
            mproject = minitab.ActiveProject
            time.sleep(1)
            mproject.SaveAs(minitabaddr_)
            return minitab, mproject
        except (PermissionError, BadZipFile, EOFError, com_error, AttributeError):
            time.sleep(1)
            continue


def minitabquit(minitab=None):
    if minitab is None:
        minitab = win32com.client.Dispatch('{D3407EDD-1BAE-4E00-A512-2673EBF849CE}')
    while 1:
        try:
            minitab.Quit()
            break
        except (PermissionError, BadZipFile, EOFError, com_error, AttributeError):
            pass


def main(phasecheck=3, signal=None):
    if signal is None:
        signal = 0
    paradifine()
    logconfi(logautorunaddr)
    if phasecheck == 3:
        rownumber_ = rownumber
        minitbookaddr__ = minitbookaddr
    elif phasecheck == 4:
        rownumber_ = rownumbersthreerounbo
        minitbookaddr__ = minitbookroundboaddr
    else:
        rownumber_ = rownumbersingle
        minitbookaddr__ = singminitbookaddr

    exc, factornumber, get_num_lista21 = runfatorn(minitbookaddr__, rownumber_)
    logdebug(get_num_lista21)
    currentlist = []
    extract_data_from_excellist(minitbookaddr__, 2, 1, 3, 9, 'Sheet', currentlist)
    try:
        minitab, mproject = minitabop(minitabaddr)
        if factornumber < 14:
            if currentlist[1] is not None and currentlist[0] is not None:
                hhcommand(rownumber_, factornumber, mproject)
            else:
                lhcommand(rownumber_, factornumber, mproject)
            if phasecheck == 3 or phasecheck == 4:
                extractdata(minitbookaddr__, signal, factornumber, rownumber_, mproject, get_num_lista21)
            else:
                extractdatasinglesmalnum(minitbookaddr__, signal, factornumber, rownumber_, mproject, get_num_lista21)

        elif factornumber < 26:

            if currentlist[1] is not None and currentlist[0] is not None:
                hhcommandb(rownumber_, factornumber, mproject)
            else:
                lhcommandb(rownumber_, factornumber, mproject)
            if phasecheck == 3 or phasecheck == 4:
                extractdata(minitbookaddr__, signal, factornumber, rownumber_, mproject, get_num_lista21)
            else:
                extractdatasinglesmalnum(minitbookaddr__, signal, factornumber, rownumber_, mproject, get_num_lista21)

        if factornumber >= 26:
            if currentlist[1] is not None and currentlist[0] is not None:
                hhcommand(rownumber_, factornumber, mproject)
            else:
                lhcommand(rownumber_, factornumber, mproject)
            if phasecheck == 4:
                extractdatathreeroundbignu(minitbookaddr__, signal, factornumber, rownumber_, mproject, get_num_lista21)
            elif phasecheck == 1:
                extractdatasinglebignu(minitbookaddr__, signal, factornumber, rownumber_, mproject, get_num_lista21)
        mproject.SaveAs(minitabaddr)
        minitabquit(minitab)

    except com_error:
        input('pywintypes.com_error:')


if __name__ == '__main__':
    main(3, 0)
