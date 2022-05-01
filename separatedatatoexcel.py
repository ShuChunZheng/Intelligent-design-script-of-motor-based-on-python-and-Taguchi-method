import shutil

import numpy
from openpyxl import Workbook

from minitabexecute import dirgenerate, rownumbermax, rownumber, write_data_to_excel, rownumbersingle, deltefile, \
    copyfile, extract_data_from_excel, rownumbersthreerounbo

columnnumber2 = 10
joffset1 = 2
global workbookaddr
global workbookaddr1
global workbookaddr2
global workbookaddr3
global workbookaddr4
global MotorCAD_Fileshot
global bigfiled
global MotorCAD_Filelist


def paradifine():
    global workbookaddr
    global workbookaddr1
    global workbookaddr2
    global workbookaddr3
    global workbookaddr4
    global MotorCAD_Fileshot
    global bigfiled
    global MotorCAD_Filelist

    diclist = dirgenerate()
    MotorCAD_Fileshot = diclist['MotorCAD_Fileshot']
    workbookaddr = diclist['workbookaddr']
    workbookaddr1 = diclist['workbookaddr1']
    workbookaddr2 = diclist['workbookaddr2']
    workbookaddr3 = diclist['workbookaddr3']
    workbookaddr4 = diclist['workbookaddr4']
    bigfiled1 = diclist['bigfiled1']
    bigfiled2 = diclist['bigfiled2']
    bigfiled3 = diclist['bigfiled3']
    bigfiled4 = diclist['bigfiled4']
    bigfiled = [bigfiled1, bigfiled2, bigfiled3, bigfiled4]
    MotorCAD_File1 = diclist['MotorCAD_File1']
    MotorCAD_File2 = diclist['MotorCAD_File2']
    MotorCAD_File3 = diclist['MotorCAD_File3']
    MotorCAD_File4 = diclist['MotorCAD_File4']
    MotorCAD_Filelist = [MotorCAD_File1, MotorCAD_File2, MotorCAD_File3, MotorCAD_File4]


def main(phasecheck=3):
    paradifine()

    workbookaddrlist = [workbookaddr1, workbookaddr2, workbookaddr3, workbookaddr4]

    if phasecheck == 3:
        rownumber_ = rownumber
    elif phasecheck == 4:
        rownumber_ = rownumbersthreerounbo
    else:
        rownumber_ = rownumbersingle

    for i in range(4):
        deltefile(workbookaddrlist[i])
        deltefile(MotorCAD_Filelist[i])

    wb1 = Workbook()
    wb2 = Workbook()
    wb3 = Workbook()
    wb4 = Workbook()

    wb1.save(workbookaddr1)
    wb2.save(workbookaddr2)
    wb3.save(workbookaddr3)
    wb4.save(workbookaddr4)

    get_num_list = numpy.zeros((rownumbermax, rownumber_))
    rownumberture = extract_data_from_excel(workbookaddr, rownumbermax, rownumber_, joffset1, get_num_list)
    if rownumberture % 4 == 0:
        boolremain = 0
    else:
        boolremain = 1
    rownumberseperate = (rownumberture // 4) + boolremain
    for i in range(rownumberture):
        if i < rownumberseperate:
            write_data_to_excel(i, i, rownumberseperate, rownumber_, workbookaddr1, get_num_list, wb1)
        elif i < rownumberseperate * 2:
            write_data_to_excel(i - rownumberseperate, i, rownumberseperate, rownumber_, workbookaddr2, get_num_list,
                                wb2)
        elif i < rownumberseperate * 3:
            write_data_to_excel(i - rownumberseperate * 2, i, rownumberseperate, rownumber_, workbookaddr3,
                                get_num_list, wb3)
        elif i < rownumberseperate * 4:
            write_data_to_excel(i - rownumberseperate * 3, i, rownumberseperate, rownumber_, workbookaddr4,
                                get_num_list, wb4)
    MotorCAD_Fileshothol = MotorCAD_Fileshot + '.mot'

    for i in range(4):
        copyfile(MotorCAD_Fileshothol, MotorCAD_Filelist[i])

    for i in bigfiled:
        try:
            shutil.rmtree(i)
        except FileNotFoundError:
            break
        except PermissionError:
            continue


if __name__ == '__main__':
    main(4)
